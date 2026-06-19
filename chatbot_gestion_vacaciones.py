from openpyxl import load_workbook


ARCHIVO_EXCEL = "DB_Gestion_Vacaciones.xlsx"


def buscar_empleado(legajo):
    wb = load_workbook(ARCHIVO_EXCEL)

    hoja = wb["Empleados"]

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] == legajo:
            wb.close()
            return {
                "legajo": fila[0],
                "nombre": fila[1],
                "dias": fila[2]
            }

    wb.close()
    return None


def fecha_bloqueada(fecha):
    wb = load_workbook(ARCHIVO_EXCEL)

    hoja = wb["Fechas_Bloqueadas"]

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if str(fila[0]) == fecha:
            wb.close()
            return True

    wb.close()
    return False


def registrar_solicitud(legajo, fecha, dias):
    wb = load_workbook(ARCHIVO_EXCEL)

    hoja = wb["Solicitudes"]

    hoja.append([
        legajo,
        fecha,
        dias,
        "Pendiente"
    ])

    wb.save(ARCHIVO_EXCEL)
    wb.close()


def solicitar_vacaciones():

    print("\n===================================")
    print(" CHATBOT - GESTION DE VACACIONES")
    print("===================================\n")

    # Estado 1
    print("Estado: Esperando Legajo")

    try:
        legajo = int(input("Ingrese su legajo: "))
    except ValueError:
        print("\nERROR: Debe ingresar un número.")
        return

    # Estado 2
    print("\nEstado: Validando Legajo")

    empleado = buscar_empleado(legajo)

    if empleado is None:
        print("\nERROR: Legajo inválido.")
        return

    print(f"\nEmpleado: {empleado['nombre']}")

    # Estado 3
    print("\nEstado: Consultando saldo de vacaciones")

    dias_disponibles = empleado["dias"]

    print(f"Días disponibles: {dias_disponibles}")

    if dias_disponibles <= 0:
        print("\nSaldo insuficiente.")
        return

    # Estado 4
    print("\nEstado: Solicitud de vacaciones")

    fecha_inicio = input(
        "Ingrese fecha de inicio (AAAA-MM-DD): "
    ).strip()

    try:
        cantidad_dias = int(
            input("Ingrese cantidad de días: ")
        )
    except ValueError:
        print("\nERROR: Debe ingresar un número.")
        return

    if cantidad_dias <= 0:
        print("\nERROR: La cantidad debe ser mayor a cero.")
        return

    if cantidad_dias > dias_disponibles:
        print(
            "\nERROR: No posee suficientes días disponibles."
        )
        return

    # Estado 5
    print("\nEstado: Validando disponibilidad")

    if fecha_bloqueada(fecha_inicio):
        print(
            "\nLas fechas seleccionadas no se encuentran disponibles."
        )
        return

    # Estado 6
    print("\nEstado: Registrando solicitud")

    registrar_solicitud(
        legajo,
        fecha_inicio,
        cantidad_dias
    )

    print("\nSolicitud registrada correctamente.")
    print("Estado: Finalizado")


def menu():

    while True:

        print("\n============================")
        print("  SISTEMA DE VACACIONES")
        print("============================")
        print("1 - Solicitar vacaciones")
        print("2 - Salir")

        opcion = input("\nSeleccione una opción: ")

        if opcion == "1":
            solicitar_vacaciones()

        elif opcion == "2":
            print("\nGracias por utilizar el sistema.")
            break

        else:
            print("\nOpción inválida.")


menu()